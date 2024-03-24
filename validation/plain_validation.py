import argparse
from torch.utils.data import DataLoader
from torchvision.transforms import Compose, ToTensor, Resize

import re
from torch.hub import tqdm
import glob
import os
import random
from UXNet_3D.networks.UXNet_3D.network_backbone import UXNET
import lightning
import numpy as np
import torch
from lightning.pytorch.loggers import TensorBoardLogger
from monai.data import pad_list_data_collate, list_data_collate, decollate_batch, DataLoader, PersistentDataset, \
    Dataset, CacheDataset
from monai.inferers import sliding_window_inference
from monai.losses import DiceLoss
from monai.metrics import DiceMetric
from monai.networks.layers import Norm
from monai.networks.nets import UNet, SwinUNETR
from monai.transforms import (
    EnsureChannelFirstd,
    AddCoordinateChannelsDict,
    AsDiscrete,
    EnsureChannelFirstd,
    Compose,
    CropForegroundd,
    LoadImaged,
    Orientationd,
    RandCropByPosNegLabeld,
    ScaleIntensityRanged,
    Spacingd,
    EnsureType, RandAffined, RandRotated, RandFlipd, Rand3DElasticd, ResizeWithPadOrCropd
)

task_name = "Task01_pancreas"
directory = os.environ.get("MONAI_DATA_DIRECTORY")
data_dir = os.path.join(directory, task_name)
persistent_cache = os.path.join(data_dir, "persistent_cache")
tensorboard_dir = os.path.join("../runs", f"{task_name}")
cuda = torch.device("cuda:0")

modulepth = r"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\runs\Task01_pancreas\UNet\version_60\checkpoints\epoch=939-step=15980.ckpt"
criterion_name = "DiceMetric"

import re

# Define the pattern using regular expression
# pattern = r"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\runs\Task01_pancreas\(\\w+)\(\\w+_\\d+)\checkpoints"
#
# # Use the search function to find the pattern in the input string
# match = re.search(pattern, modulepth)
#
# # Check if a match is found
# assert match, "module pth do not contain the pattern"
    # Extract the matched groups
# network_name = match.group(1)
# version_number = match.group(2)

network_name = r"SwinUNETR"
version_number = r"version_6"

args = argparse.Namespace(
    no_cuda=False,  # disables CUDA training
    epochs=10,  # number of epochs (default : 10)
    lr=1e-2,  # base learning rate (default : 0.01)
    weight_decay=3e-2,  # weight decay value (default : 0.03)
    batch_size=1,  # batch size (default : 4)
    dry_run=False  # quickly check a single pass
)

block_list = ['0006', '0009', '0013', '0022', '0027', '0032', '0048', '0051',
              '0078', '0079', '0082', '0088', '0089', '0094', '0100', '0113',
              '0131', '0136', '0139', '0142']


class PlainVal:

    def __init__(self, args, model, val_dataloader, criterion, device):
        self.model = model

        self.val_dataloader = val_dataloader
        self.criterion = criterion
        self.epoch = args.epochs
        self.device = device
        self.args = args
        self.post_pred = Compose(
            [EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(argmax=True, to_onehot=2)])
        self.post_label = Compose([EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(to_onehot=2)])

    def eval_fn(self, current_epoch):
        self.model.eval()
        total_loss = 0.0
        tk = tqdm(self.val_dataloader, desc="EPOCH" + "[VALID]" + str(current_epoch + 1) + "/" + str(self.epoch))

        with torch.no_grad():
            for t, data in enumerate(tk):
                images, labels = data["image"], data["label"]
                images, labels = images.to(self.device), labels.to(self.device)

                # logits = sliding_window_inference(images, roi_size=[60,60,60], sw_batch_size=1, predictor=self.model)
                logits = self.model(images)
                logits = [self.post_pred(i) for i in decollate_batch(logits)]
                labels = [self.post_label(i) for i in decollate_batch(labels)]
                self.criterion(logits, labels)
                pass

    def valdation(self):
        best_dice_metric = np.inf

        for i in range(self.epoch):
                self.eval_fn(i)
                dice_metric = self.criterion.aggregate().item()
                all.append(dice_metric)
                print(f"Dice Metric : {dice_metric}")
        self.criterion.reset()

def main():
    use_cuda = not args.no_cuda and torch.cuda.is_available()
    device = torch.device("cuda" if use_cuda else "cpu")

    # 读数据的名字
    train_images = sorted(glob.glob(os.path.join(data_dir, "img_proc", "*.nii.gz")))
    train_labels = sorted(glob.glob(os.path.join(data_dir, "pancreas_seg_proc", "*.nii.gz")))

    fltr_train_images = [string for string in train_images if not any(block in string for block in block_list)]
    fltr_train_labels = [string for string in train_labels if not any(block in string for block in block_list)]

    data_dicts = [{"image": image_name, "label": label_name} for image_name, label_name in
                  zip(fltr_train_images, fltr_train_labels)]  # 注意到data_dicts是一个数组

    data_dicts = random.sample(data_dicts, round(0.25 * len(data_dicts)))
    # 拆分成train和val
    # validation_files = random.sample(data_dicts, round(0.3 * len(data_dicts)))

    # 挑数据集，把整个data_dicts都归进validation_files
    validation_files = data_dicts

    val_transforms = Compose(
        [
            LoadImaged(keys=["image", "label"]),
            EnsureChannelFirstd(keys=["image", "label"]),
            RandCropByPosNegLabeld(
                keys=["image", "label"],
                label_key="label",
                spatial_size=(96, 96, 96),
                pos=1,
                neg=1,
                num_samples=4,
                image_key="image",
                image_threshold=0,
                # allow_smaller=True
            ),
            ResizeWithPadOrCropd(keys=["image", "label"], spatial_size=(96, 96, 96))
        ]
    )

    validation_dataset = CacheDataset(data=validation_files, transform=val_transforms)
    valid_loader = DataLoader(validation_dataset, batch_size=args.batch_size, shuffle=True,
                              collate_fn=pad_list_data_collate, num_workers=4)

    # model = UNet(
    #     spatial_dims=3,
    #     in_channels=1,
    #     out_channels=2,
    #     channels=(16, 32, 64, 128, 256),
    #     strides=(2, 2, 2, 2),
    #     num_res_units=2,
    #     norm=Norm.BATCH,
    # )

    # model = UXNET(
    #     in_chans=1,
    #     out_chans=2,
    #     depths=[2, 2, 2, 2],
    #     feat_size=[48, 96, 192, 384],
    #     drop_path_rate=0,
    #     layer_scale_init_value=1e-6,
    #     spatial_dims=3,
    # ).to(device)

    model = SwinUNETR(
            spatial_dims=3,
            in_channels=1,
            out_channels=2,
            img_size=(96, 96, 96)
        ).to(device)

    ckpt = torch.load(r"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\runs\Task01_pancreas\SwinUNETR\version_16\checkpoints\epoch=24-step=1800.ckpt")

    new_state_dict = {}
    for key, value in ckpt["state_dict"].items():
        if key.startswith("_model."):
            new_key = key[len("_model."):]  # Remove the "_module" prefix
            new_state_dict[new_key] = value
        else:
            new_state_dict[key] = value

    new_state_dict.pop("loss_function.class_weight")
    model.load_state_dict(new_state_dict)
    model.to(device)

    criterion = DiceMetric(include_background=False, reduction="mean", get_not_nans=False)

    PlainVal(args, model, valid_loader, criterion, device).valdation()

def write_record():
        now = datetime.datetime.now()
        instant_time = now.strftime("%Y-%m-%d\n"
                                    "%H:%M:%S")
        wb = load_workbook(fr"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\validation_results\{network_name}.xlsx")

        # Check if the sheet exists
        if network_name in wb.sheetnames:
            # If the sheet exists, select it
            sheet = wb[network_name]
        else:
            # If the sheet does not exist, create it
            sheet = wb.create_sheet(network_name)
        ws = wb.active

        column = 1
        while ws.cell(1, column).value is not None:
            column += 2
        # Iterate over the list and append each item to the worksheet
        sheet.cell(row=1, column=column, value=instant_time)
        sheet.cell(row=2, column=column, value=criterion_name)

        for i, mean_dice in enumerate(all):
            sheet.cell(row=i + 3, column=column,
                       value=mean_dice)  # append() expects a list, even if it's a single element
        # Save the workbook to a new xlsx file
        sheet.cell(row=len(all)+3, column=column, value=np.mean(all))
        wb.save(fr"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\validation_results\{network_name}.xlsx")
        wb.close()


if __name__ == "__main__":
    os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:512"

    all=[]
    from openpyxl import load_workbook
    import datetime
    import atexit
    atexit.register(write_record)

    main()

