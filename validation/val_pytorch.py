import argparse
import glob
import os
import random
from datetime import datetime

import numpy as np
import torch
from torch.hub import tqdm
from torch.utils.data import DataLoader
from torch.utils.tensorboard import SummaryWriter
from torchvision.transforms import Compose

from monai.data import pad_list_data_collate, decollate_batch, DataLoader, CacheDataset
from monai.losses import DiceLoss
from monai.metrics import DiceMetric
from monai.networks.nets import SwinUNETR
from monai.transforms import (
    Compose, LoadImaged, RandCropByPosNegLabeld,
    RandAffined, RandRotated, RandFlipd, ResizeWithPadOrCropd, EnsureChannelFirstd, EnsureType, AsDiscrete
)
from my_network.SwinViT_Upp import SwinViT_Upp
from my_network.SwinViT_Upp_attg import SwinViT_Upp_attg

from openpyxl import load_workbook
import datetime
import atexit

from my_network.SwinViT_uxblock_Upp import SwinViT_uxblock_Upp

args = argparse.Namespace(
    task_name="Task01_pancreas",
    network_name="SwinViT_Upp_pytorch",
    criterion_name="DiceMetric",
    load_model_path=r"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\runs_nocheat\Task01_pancreas\SwinViT_uxblock_Upp_loss\0404_2342\checkpoint\best130_0.77473.pth",
    epochs=3,  # number of epochs (default : 10)
    batch_size=1,  # batch size (default : 4)
    dry_run=False  # quickly check a single pass
)


directory = os.environ.get("MONAI_DATA_DIRECTORY")
data_dir = os.path.join(directory, args.task_name)
persistent_cache = os.path.join(data_dir, "persistent_cache")
cuda = torch.device("cuda:0")

all=[]

class TrainEval:

    def __init__(self, args, model, val_dataloader, metric, device):
        # 训练过程中的指标
        self.val_metric_latest = 0  # 上一次val的metric
        self.hvm_epoch = 0  # 取得highest val metric对应的epoch
        self.highest_val_metric = -np.inf  # 历次val的最高metric
        self.model = model

        self.val_dataloader = val_dataloader
        self.metric = metric  # val用的metric指标（DSC）


        self.device = device
        self.args = args  # 超参数包

        # 我觉得需要后处理，但是没有好像也能跑
        self.post_pred = Compose(
            [EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(argmax=True, to_onehot=2)])
        self.post_label = Compose([EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(to_onehot=2)])

    def eval_fn(self, current_epoch):
        self.model.eval()
        tk_step = tqdm(self.val_dataloader, desc="EPOCH" + "[VALID]" + str(current_epoch),
                       position=0, leave=True, dynamic_ncols=True)  # Added dynamic_ncols=True

        with torch.no_grad():
            for t, data in enumerate(tk_step):
                images, labels = data["image"], data["label"]
                images, labels = images.to(self.device), labels.to(self.device)

                logits = self.model(images)
                logits = logits[4]
                logits = [self.post_pred(i) for i in decollate_batch(logits)]  # 这里用list加后处理没问题
                labels = [self.post_label(i) for i in decollate_batch(labels)]  # 可以运行
                self.metric(logits, labels)
                # 这里好像是可以用的，先不改
                tk_step.set_postfix({"metric": "%4f" % round(self.metric.aggregate().item(), 5),
                                     "hvm": f"{round(self.highest_val_metric, 5)}"})

            metric = self.metric.aggregate().item()
            self.metric.reset()
        return metric

    def val(self):
        td = tqdm(range(self.args.epochs), desc="val",
                  position=0, leave=True, dynamic_ncols=True)  # Added dynamic_ncols=True
        # epochs循环
        for i in td:
            current_epoch = i + 1
            self.val_metric_latest = self.eval_fn(current_epoch)  # 赋给全局变量保存下来
            print(f"val_metric{current_epoch}={round(self.val_metric_latest, 5)}")
            all.append(self.val_metric_latest)

            # 如果新最高，保存这个metric和对应epoch
            if self.val_metric_latest > self.highest_val_metric:
                self.highest_val_metric = self.val_metric_latest
                self.hvm_epoch = current_epoch


def main():
    # writer和checkpoint路径必须要在main()里面声明，我是真的不知道为什么放在里面就没事
    # 放在外面会不断地生成新的文件夹和新的event文件，到底为什么
    # 放在外面是文件作用域，整个文件里的所有调用都直接运行它，所以会生成多个instance，反复创建新文件夹
    # 放在里面是函数作用域，只有调用这个函数时才运行一次，函数没有推出，调用的就一直是同一个instance
    torch.set_float32_matmul_precision("medium")
    os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:512"

    # use_cuda = not args.no_cuda and torch.cuda.is_available()
    # device = torch.device("cuda" if use_cuda else "cpu")
    device = torch.device("cuda:0")

    # 读数据的名字
    test_images = sorted(glob.glob(os.path.join(data_dir, "img_proc", "test", "*.nii.gz")))
    test_labels = sorted(glob.glob(os.path.join(data_dir, "pancreas_seg_proc", "test", "*.nii.gz")))
    test_dicts = [{"image": image_name, "label": label_name} for image_name, label_name in
                  zip(test_images, test_labels)]

    val_transforms = Compose(
        [
            LoadImaged(keys=["image", "label"]),
            EnsureChannelFirstd(keys=["image", "label"]),
            RandCropByPosNegLabeld(
                keys=["image", "label"],
                label_key="label",
                spatial_size=(64, 64, 64),
                pos=3,
                neg=1,
                num_samples=4,
                image_key="image",
                image_threshold=0,
                allow_smaller=True
            ),
            ResizeWithPadOrCropd(keys=["image", "label"], spatial_size=(64, 64, 64))
        ]
    )

    valid_dataset = CacheDataset(data=test_dicts, transform=val_transforms)
    valid_loader = DataLoader(valid_dataset, batch_size=args.batch_size, shuffle=True, num_workers=4,
                              collate_fn=pad_list_data_collate)

    # model = UNet(
    #         spatial_dims=3,
    #         in_channels=1,
    #         out_channels=2,
    #         channels=(16, 32, 64, 128, 256),
    #         strides=(2, 2, 2, 2),
    #         num_res_units=2,
    #         norm=Norm.BATCH,
    #     ).to(device)

    # model = UXNET(
    #     in_chans=1,
    #     out_chans=2,
    #     depths=[2, 2, 2, 2],
    #     feat_size=[48, 96, 192, 384],
    #     drop_path_rate=0,
    #     layer_scale_init_value=1e-6,
    #     spatial_dims=3,
    # ).to(device)

    # model = SwinUNETR(
    #     spatial_dims=3,
    #     in_channels=1,
    #     out_channels=2,
    #     img_size=(64, 64, 64)
    # )

    model = SwinViT_uxblock_Upp()

    ckpt = torch.load(args.load_model_path)
    # lightning训练出的模型
    if "state_dict" in ckpt:
        new_state_dict = {}
        for key, value in ckpt["state_dict"].items():
            if key.startswith("_model."):
                new_key = key[len("_model."):]  # Remove the "_module" prefix
                new_state_dict[new_key] = value
            else:
                new_state_dict[key] = value
        new_state_dict.pop("loss_function.class_weight")
        model.load_state_dict(new_state_dict)
    # pytorch训练的模型
    else:
        model.load_state_dict(ckpt)
    model.to(device)

    metric = DiceMetric(include_background=False, reduction="mean")

    TrainEval(args, model, valid_loader, metric,device).val()

def write_record():
    now = datetime.datetime.now()
    instant_time = now.strftime("%m%d-%H%M")
    # 所有数据写到一个excel文件下，用sheet的名字区分不同的网络
    wb = load_workbook(fr"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\validation_results\val_result.xlsx")

    # 有对应的sheet就打开
    if args.network_name in wb.sheetnames:
        sheet = wb[args.network_name]
    else:
        # 没有这个网络命名的sheet就现创一个
        sheet = wb.create_sheet(args.network_name)
    ws = wb.active

    # openyxl的行和列编号是从1开始的
    column = 1
    while ws.cell(1, column).value is not None:
        column += 2
    # Iterate over the list and append each item to the worksheet
    sheet.cell(row=1, column=column, value=instant_time)
    sheet.cell(row=2, column=column, value=args.criterion_name)

    for i, mean_dice in enumerate(all):
        sheet.cell(row=i + 3, column=column,
                   value=mean_dice)  # append() expects a list, even if it's a single element
    # Save the workbook to a new xlsx file
    sheet.cell(row=len(all) + 3, column=column, value=np.mean(all))
    wb.save(fr"C:\Git\NeuralNetwork\Pancreas_with_Monai_Lightning\validation_results\val_result.xlsx")
    wb.close()

if __name__ == "__main__":
    main()
    write_record()
