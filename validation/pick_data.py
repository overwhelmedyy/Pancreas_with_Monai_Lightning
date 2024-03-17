import argparse
import re
from torch.hub import tqdm
import glob
import os
import torch
from monai.data import pad_list_data_collate, list_data_collate, decollate_batch, DataLoader, PersistentDataset, \
    Dataset, CacheDataset
from monai.metrics import DiceMetric
from monai.networks.layers import Norm
from monai.networks.nets import UNet
from monai.transforms import (
    AsDiscrete,
    EnsureChannelFirstd,
    Compose,
    CropForegroundd,
    LoadImaged,
    Orientationd,
    RandCropByPosNegLabeld,
    ScaleIntensityRanged,
    Spacingd,
    EnsureType,
    ResizeWithPadOrCropd
)

task_name = "Task01_pancreas"
network_name = "UNet"

directory = os.environ.get("MONAI_DATA_DIRECTORY")
data_dir = os.path.join(directory, task_name)
persistent_cache = os.path.join(data_dir, "persistent_cache")
tensorboard_dir = os.path.join("../runs", f"{task_name}")
cuda = torch.device("cuda:0")

args = argparse.Namespace(
    no_cuda=False,  # disables CUDA training
    epochs=1,  # number of epochs (default : 10)
    lr=1e-2,  # base learning rate (default : 0.01)
    weight_decay=3e-2,  # weight decay value (default : 0.03)
    batch_size=1,  # batch size (default : 4)
    dry_run=False  # quickly check a single pass
)

import torch
from openpyxl import Workbook

# Create a new workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = 'sheet1'



if __name__ == "__main__":
    os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:512"

    val_transforms = Compose(
        [
            LoadImaged(keys=["image", "label"]),
            EnsureChannelFirstd(keys=["image", "label"]),
            # LabelToMaskd(keys=['label'], select_labels=[0,2]),
            Orientationd(keys=["image", "label"], axcodes="RAS"),
            Spacingd(
                keys=["image", "label"],
                pixdim=(1.5, 1.5, 2.0),
                mode=("bilinear", "nearest"),
            ),
            ScaleIntensityRanged(
                keys=["image"],
                a_min=-1024,
                a_max=2976,
                b_min=0.0,
                b_max=1.0,
                clip=True,
            ),
            CropForegroundd(keys=["image", "label"], source_key="image"),
            RandCropByPosNegLabeld(
                keys=["image", "label"],
                label_key="label",
                spatial_size=(160, 160, 160),
                pos=1,
                neg=1,
                num_samples=20,
                image_key="image",
                image_threshold=0,
                allow_smaller=True
            ),
            ResizeWithPadOrCropd(keys=["image", "label"], spatial_size=(160, 160, 160))
        ]
    )

    use_cuda = not args.no_cuda and torch.cuda.is_available()
    device = torch.device("cuda" if use_cuda else "cpu")

    model = UNet(
        spatial_dims=3,
        in_channels=1,
        out_channels=2,
        channels=(16, 32, 64, 128, 256),
        strides=(2, 2, 2, 2),
        num_res_units=2,
        norm=Norm.BATCH,
    )
    ckpt = torch.load("../runs/Task01_pancreas/UNet/version_60/checkpoints/epoch=939-step=15980.ckpt")

    post_pred = Compose(
        [EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(argmax=True, to_onehot=2)])
    post_label = Compose([EnsureType("tensor", device=torch.device("cpu")), AsDiscrete(to_onehot=2)])

    new_state_dict = {}
    for key, value in ckpt["state_dict"].items():
        if key.startswith("_model."):
            new_key = key[len("_model."):]  # Remove the "_module" prefix
            new_state_dict[new_key] = value
        else:
            new_state_dict[key] = value

    new_state_dict.pop("loss_function.class_weight")
    model.load_state_dict(new_state_dict)
    model.to(device)
    criterion = DiceMetric(include_background=False, reduction="mean", get_not_nans=False)

    # 读数据的名字
    train_images = sorted(glob.glob(os.path.join(data_dir, "img", "*.nii.gz")))
    train_labels = sorted(glob.glob(os.path.join(data_dir, "pancreas_seg", "*.nii.gz")))
    data_dicts = [{"image": image_name, "label": label_name} for image_name, label_name in
                  zip(train_images, train_labels)]  # 注意到data_dicts是一个数组

    tk_image = tqdm(data_dicts, desc="EPOCH" + "[VALID]" + str(1) + "/" + str(len(data_dicts)))

    pattern = r'(\d{4})\.nii\.gz'
    for row, image in enumerate(tk_image):
          # 第一层 把图片单张解出来
        validation_files = [image]
        validation_dataset = PersistentDataset(data=validation_files, transform=val_transforms,
                                               cache_dir=persistent_cache)

        matches = re.findall(pattern, validation_files[0]['image'])
        if matches:
            extracted_number = matches[0]
            ws.cell(row+2, 1, extracted_number) # row+1是把第一行表头空出来
        else:
            RuntimeError("No match found")

        for column in tqdm(range(10)):
            valid_loader = DataLoader(validation_dataset, batch_size=4, shuffle=False,
                                      collate_fn=pad_list_data_collate, num_workers=4)
            model.eval()

            tk_eval_fn = tqdm(valid_loader,desc="EPOCH" + "[VALID]" + str(row + 1) + "/" + str(len(data_dicts)))

            for _, data in enumerate(tk_eval_fn):
                images, labels = data["image"], data["label"]
                images, labels = images.to(device), labels.to(device)
                logits = model(images)
                logits = [post_pred(i) for i in decollate_batch(logits)]
                labels = [post_label(i) for i in decollate_batch(labels)]
                criterion(logits, labels)
            dice_metric = criterion.aggregate().item()
            ws.cell(row+2, column+2, dice_metric)
            criterion.reset()
            wb.save(r'./logs/example.xlsx')



